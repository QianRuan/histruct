#encoding=utf-8


import argparse
from datetime import datetime
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THICK, BORDER_THIN
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from others.logging import init_logger,logger

val_xents_file = 'validation_xent.json'
test_xents_file = 'test_xent.json'
test_rouges_file = 'test_rouges.json'
test_avg_rouges_file = 'test_avg_rouges.json'
 
metrics=['rouge_1_f_score','rouge_2_f_score','rouge_l_f_score',
                 'rouge_1_recall','rouge_2_recall','rouge_l_recall',
                 'rouge_1_precision','rouge_2_precision','rouge_l_precision']

cnndm_reported_baselines = []
cnndm_reported_baselines.append({'model':'MatchSum(RoBERTa-base)', 'rouge_1_f_score':44.41, 'rouge_2_f_score':20.86, 'rouge_l_f_score':40.55})
cnndm_reported_baselines.append({'model':'MatchSum(BERT-base)', 'rouge_1_f_score':44.22, 'rouge_2_f_score':20.62, 'rouge_l_f_score':40.38})
cnndm_reported_baselines.append({'model':'BERTSUMEXT', 'rouge_1_f_score':43.25, 'rouge_2_f_score':20.24, 'rouge_l_f_score':39.63})
cnndm_reported_baselines.append({'model':'BERTSUMEXT w/o interval embeddings', 'rouge_1_f_score':43.20, 'rouge_2_f_score':20.22, 'rouge_l_f_score':39.59})
cnndm_reported_baselines.append({'model':'BERTSUMEXT (large)', 'rouge_1_f_score':43.85, 'rouge_2_f_score':20.34, 'rouge_l_f_score':39.90})
cnndm_reported_baselines.append({'model':'BERTSUM-lead3', 'rouge_1_f_score':40.42, 'rouge_2_f_score':17.62, 'rouge_l_f_score':36.67})
cnndm_reported_baselines.append({'model':'BERTSUM-oracle', 'rouge_1_f_score':52.59, 'rouge_2_f_score':31.24, 'rouge_l_f_score':48.87})

pubmed_reported_baselines = []
pubmed_reported_baselines.append({'model':'HAT-BART', 'rouge_1_f_score':48.25, 'rouge_2_f_score':21.35, 'rouge_l_f_score':36.69})
pubmed_reported_baselines.append({'model':'DANCER PEAGASUS', 'rouge_1_f_score':46.34, 'rouge_2_f_score':19.97, 'rouge_l_f_score':42.42})
pubmed_reported_baselines.append({'model':'BigBird-Pegasus', 'rouge_1_f_score':46.32, 'rouge_2_f_score':20.65, 'rouge_l_f_score':42.33})
pubmed_reported_baselines.append({'model':'ExtSum-LG+MMR-Select+', 'rouge_1_f_score':45.39, 'rouge_2_f_score':20.37, 'rouge_l_f_score':40.99})
pubmed_reported_baselines.append({'model':'ExtSum-LG+RdLoss', 'rouge_1_f_score':45.3, 'rouge_2_f_score':20.42, 'rouge_l_f_score':40.95})
pubmed_reported_baselines.append({'model':'PEGASUS', 'rouge_1_f_score':45.49, 'rouge_2_f_score':19.90, 'rouge_l_f_score':42.42})

arxiv_reported_baselines = []
arxiv_reported_baselines.append({'model':'HAT-BART', 'rouge_1_f_score':46.74, 'rouge_2_f_score':19.19, 'rouge_l_f_score':42.2})
arxiv_reported_baselines.append({'model':'LED-large (seqlen: 16,384)', 'rouge_1_f_score':46.63, 'rouge_2_f_score':19.62, 'rouge_l_f_score':41.48})
arxiv_reported_baselines.append({'model':'BigBird-Pegasus', 'rouge_1_f_score':46.63, 'rouge_2_f_score':19.02, 'rouge_l_f_score':41.77})
arxiv_reported_baselines.append({'model':'DANCER PEAGASUS', 'rouge_1_f_score':45.01, 'rouge_2_f_score':17.6, 'rouge_l_f_score':40.56})
arxiv_reported_baselines.append({'model':'PEGASUS', 'rouge_1_f_score':44.70, 'rouge_2_f_score':17.27, 'rouge_l_f_score':25.80})
arxiv_reported_baselines.append({'model':'ExtSum-LG+RdLoss', 'rouge_1_f_score':44.01, 'rouge_2_f_score':17.79, 'rouge_l_f_score':39.09})
arxiv_reported_baselines.append({'model':'ExtSum-LG+MMR-Select+', 'rouge_1_f_score':43.87, 'rouge_2_f_score':17.5, 'rouge_l_f_score':38.97})



def str2bool(v):
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')


def get_test_rouges_result(model,eval_path):
    
    test_rouges_path = eval_path + test_rouges_file
    test_avg_rouges_path = eval_path + test_avg_rouges_file
        
    #test avg. rouges
    test_avg_rouges_result = {'model':None}
    for m in metrics:
        test_avg_rouges_result.update({m:None})
        
    with open(test_avg_rouges_path, 'r') as f:
        test_avg_rouges = json.load(f)
        
    test_avg_rouges_result['model']=model
    for m in metrics:
        test_avg_rouges_result[m]= round(test_avg_rouges[m] * 100, 2)
    
    # test rouges of top 3 steps of the model
    with open(test_rouges_path, 'r') as f:
        test_rouges = json.load(f)
    
    test_rouges_results =[]
    
    if not (model.split('_')[1]=='oracle' or model.split('_')[1].startswith('lead')):
        for r in test_rouges:
            
            test_rouges_result = {'model':None, 'step':None}
            for m in metrics:
                test_rouges_result.update({m:None})
                
            test_rouges_result['model']=model
          
            test_rouges_result['step']=r[0].split('_')[-1].split('.')[0]
            
            for m in metrics:
                test_rouges_result[m]= round(r[1][m] * 100, 2)
                
            test_rouges_results.append(test_rouges_result)
    else:
        test_rouges_result = {'model':None, 'step':None}
        for m in metrics:
            test_rouges_result.update({m:None})
        
        test_rouges_result['model']=model   
        for m in metrics:
            test_rouges_result[m]= round(test_rouges[m] * 100, 2) 
        test_rouges_results.append(test_rouges_result)
      
    return test_avg_rouges_result, test_rouges_results
        
def get_rouges_df(models):
    
    df_cols = ['model'] + metrics
    df_cols1 = ['model','step'] + metrics
       
    avg_rows = []
    step_rows = []
    
    for model in models:
        
        eval_path = args.models_path + model + '/eval/'
    
        if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
            
            test_avg_rouges_result, test_rouges_results = get_test_rouges_result(model, eval_path)
            
            avg_rows.append(test_avg_rouges_result)
            step_rows = step_rows + test_rouges_results
        
        else:
            if not os.path.exists(args.models_path + model+'/DONE'):
                logger.info("---Training of the model is not finished, skip it-------%s"%(model))
            if not os.path.exists(eval_path+'DONE'):
                logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
                
                
    df1 = pd.DataFrame(avg_rows, columns = df_cols) 
    df2 = pd.DataFrame(step_rows, columns = df_cols1)                

    return df1, df2

def check_best_models(df):
    
    best_models = {}
    for m in metrics:
        best_models.update({m:[]})
    
    
        
    for m in metrics:
        max_v = df[m].max()
        for index, row in df.iterrows():
            if row[m] == max_v:
                model = row['model']
                if 'step' in df.columns:
                    step = row['step']
                else:
                    step=None
                best_models[m].append((model, step))
  
    return best_models
        
        
 
def save_eval_results_to_excel(excelfile,sheetname,df):
   
    if not os.path.isfile(excelfile):
        logger.info('The excel file is not existing, creating a new excel file...'+excelfile)       
        wb = Workbook()
        wb.save(excelfile)
        
        
    wb = load_workbook(excelfile)
    if not (sheetname in wb.sheetnames):
        logger.info('The worksheet is not existing, creating a new worksheet...'+sheetname) 
        ws1 = wb.create_sheet(sheetname)
        ws1.title = sheetname
        wb.save(excelfile)
    
            
    book = load_workbook(excelfile)
    idx = wb.sheetnames.index(sheetname)
    ws=book.get_sheet_by_name(sheetname)
    book.remove(ws)
    book.create_sheet(sheetname, idx)
    writer = pd.ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    df.to_excel(writer,sheet_name=sheetname, index = False,header= True)
    writer.save()

def color_the_best_metric(excelfile, sheetname, best_models, color, font):
    wb = load_workbook(excelfile)
    ws = wb[sheetname]
   
    # Create a dictionary of column names
    ColNames = {}
    Current  = 0
    for COL in ws.iter_cols(1, ws.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
 
    # Color best metrics 
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for m in metrics:
            bests = best_models[m]
            for model in bests:
                if 'step' in ColNames.keys():
                    if row_cells[ColNames['model']].value == model[0] and row_cells[ColNames['step']].value == model[1]:
                        
                        row_cells[ColNames[m]].fill = PatternFill("solid", fgColor=color)
                        row_cells[ColNames[m]].font = Font(b=font)
                else:
                    if row_cells[ColNames['model']].value == model[0]:
                        
                        row_cells[ColNames[m]].fill = PatternFill("solid", fgColor=color)
                        row_cells[ColNames[m]].font = Font(b=font)
                     
    wb.save(excelfile)
    
def color_the_best_metric_border(excelfile, sheetname, best_models, color):
    wb = load_workbook(excelfile)
    ws = wb[sheetname]
    
    thick_border = Border(
    left=Side(border_style=BORDER_THICK, color=color),
    right=Side(border_style=BORDER_THICK, color=color),
    top=Side(border_style=BORDER_THICK, color=color),
    bottom=Side(border_style=BORDER_THICK, color=color)
)
    thin_border = Border(
    left=Side(border_style=BORDER_THIN, color=color),
    right=Side(border_style=BORDER_THIN, color=color),
    top=Side(border_style=BORDER_THIN, color=color),
    bottom=Side(border_style=BORDER_THIN, color=color)
)
   
    # Create a dictionary of column names
    ColNames = {}
    Current  = 0
    for COL in ws.iter_cols(1, ws.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
 
    # border best metrics 
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for m in metrics:
            bests = best_models[m]
            for model in bests:
                if 'step' in ColNames.keys():
                    if row_cells[ColNames['model']].value == model[0] and row_cells[ColNames['step']].value == model[1]:
                        
                        row_cells[ColNames[m]].border = thin_border
                        
                else:
                    if row_cells[ColNames['model']].value == model[0]:
                        
                        row_cells[ColNames[m]].border = thin_border
                     
    wb.save(excelfile)
            

    
def copy_result_file(source):
    
    #copy the excelfile  
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M")
    dir_path = source[:-5]+'_Copy'
    if not os.path.isdir(dir_path):
     os.mkdir(dir_path)
    target = dir_path +'/Copy'+dt_string +'.xlsx'
    wb = load_workbook(source)
    wb.save(target)
    
    return target

def mark_best_models(best_models, df):
    if 'step' not in df.columns:
        for m in metrics:
            bests = best_models[m]
            for model in bests:
                for i in range(df.shape[0]):
                    if model[0] == df['model'].values[i].split('!')[-1]:
                         df['model'].values[i] = '!' + df['model'].values[i]
    else:
        for m in metrics:
            bests = best_models[m]
            for model in bests:
                for i in range(df.shape[0]):
                    if model[0] == df['model'].values[i].split('!')[-1] and str(model[1]) == df['step'].values[i]:
                         df['model'].values[i] ='!' + df['model'].values[i]
      
def generate_eval_results_overview(args):
    logger.info("=================================================")
    logger.info("Generating evaluation results overview...")
    
    models = sorted(os.listdir(args.models_path))
    models = [model for model in models if model.startswith(args.dataset+'_')]
    
    baseline_models = [model for model in models if model.split('_')[1]!='hs']  #or model.split('_')[1]=='oracle' or model.split('_')[1].startswith('lead')]
    baseline_models.reverse()
    baseline_models2 = [model for model in baseline_models if model.split('_')[1]!='oracle' and not model.split('_')[1].startswith('lead')]
    baseline_bert_base_models = [model for model in baseline_models if model.split('_')[1]=='bert']
    baseline_bert_large_models = [model for model in baseline_models if model.split('_')[1]=='bertL']
    baseline_roberta_base_models = [model for model in baseline_models if model.split('_')[1]=='robertaB']
    baseline_roberta_large_models = [model for model in baseline_models if model.split('_')[1]=='robertaL']
    baseline_longformer_base_models = [model for model in baseline_models if model.split('_')[1]=='longformerB']
    baseline_longformer_large_models = [model for model in baseline_models if model.split('_')[1]=='longformerL']
    baseline_bp_models = [model for model in baseline_models if model.split('_')[1]=='bigbirdP']
    baseline_bart_base_models = [model for model in baseline_models if model.split('_')[1]=='bartB']
    baseline_bart_large_models = [model for model in baseline_models if model.split('_')[1]=='bartL']
    
    
    
    histruct_models = [model for model in models if model.split('_')[1]=='hs']
    histruct_bert_base_models = [model for model in histruct_models if model.split('_')[2]=='bert']
    histruct_bert_large_models = [model for model in histruct_models if model.split('_')[2]=='bertL']
    histruct_roberta_base_models = [model for model in histruct_models if model.split('_')[2]=='robertaB']
    histruct_roberta_large_models = [model for model in histruct_models if model.split('_')[2]=='robertaL']
    histruct_longformer_base_models = [model for model in histruct_models if model.split('_')[2]=='longformerB']
    histruct_longformer_large_models = [model for model in histruct_models if model.split('_')[2]=='longformerL']
    histruct_bp_models = [model for model in histruct_models if model.split('_')[2]=='bigbirdP']
    histruct_bart_base_models = [model for model in histruct_models if model.split('_')[2]=='bartB']
    histruct_bart_large_models = [model for model in histruct_models if model.split('_')[2]=='bartL']
    
    logger.info("DATASET: %s"%(args.dataset))
    logger.info("There are %i baseline models"%(len(baseline_models)))
    logger.info("There are %i LM baseline models"%(len(baseline_models2)))
    logger.info("There are %i histruct models"%(len(histruct_models)))
    logger.info("--------- %i histruct bert_base models"%(len(histruct_bert_base_models)))
    logger.info("--------- %i histruct bert_large models"%(len(histruct_bert_large_models)))
    logger.info("--------- %i histruct roberta_base models"%(len(histruct_roberta_base_models)))
    logger.info("--------- %i histruct roberta_large models"%(len(histruct_roberta_large_models)))
    logger.info("--------- %i histruct bart_base models"%(len(histruct_bart_base_models)))
    logger.info("--------- %i histruct bart_large models"%(len(histruct_bart_large_models)))
    logger.info("--------- %i histruct longformer_base models"%(len(histruct_longformer_base_models)))
    logger.info("--------- %i histruct longformer_large models"%(len(histruct_longformer_large_models)))
    logger.info("--------- %i histruct bigbird-pegasus models"%(len(histruct_bp_models)))
    
    
    
    
    df1, df2 = get_rouges_df(baseline_models)
    df3, df4 = get_rouges_df(histruct_models)
    df3_1, df4_1 = get_rouges_df(histruct_bert_base_models)
    df3_2, df4_2 = get_rouges_df(histruct_bert_large_models)
    df3_3, df4_3 = get_rouges_df(histruct_longformer_base_models)
    df3_4, df4_4 = get_rouges_df(histruct_longformer_large_models)
    df3_5, df4_5 = get_rouges_df(histruct_bp_models)
    df3_6, df4_6 = get_rouges_df(histruct_roberta_base_models)
    df3_7, df4_7 = get_rouges_df(histruct_roberta_large_models)
    df3_8, df4_8 = get_rouges_df(histruct_bart_base_models)
    df3_9, df4_9 = get_rouges_df(histruct_bart_large_models)
    
  
    df_cols = ['model'] + metrics
    df_cols1 = ['model','step'] + metrics
    
    if args.dataset=='cnndm':
        reported_baselines=cnndm_reported_baselines
    elif args.dataset=='pubmed':
        reported_baselines=pubmed_reported_baselines
    elif args.dataset=='arxiv':
        reported_baselines=arxiv_reported_baselines
    df5 = pd.DataFrame(reported_baselines, columns = df_cols) 
    df6 = pd.DataFrame(reported_baselines, columns = df_cols1) 
    
    df7 = pd.DataFrame([{'model':'REPORTED BASELINES------------'}], columns = df_cols) 
    df8 = pd.DataFrame([{'model':'BASELINES------------'}], columns = df_cols) 
    df9 = pd.DataFrame([{'model':'OUR MODELS------------'}], columns = df_cols) 
    df9_1 = pd.DataFrame([{'model':'------------bert_base models------------'}], columns = df_cols) 
    df9_2 = pd.DataFrame([{'model':'------------bert_large models------------'}], columns = df_cols) 
    df9_3 = pd.DataFrame([{'model':'------------longformer_base models------------'}], columns = df_cols) 
    df9_4 = pd.DataFrame([{'model':'------------longformer_large models------------'}], columns = df_cols) 
    df9_5 = pd.DataFrame([{'model':'------------bigbird-pegasus models------------'}], columns = df_cols) 
    df9_6 = pd.DataFrame([{'model':'------------roberta_base models------------'}], columns = df_cols) 
    df9_7 = pd.DataFrame([{'model':'------------roberta_large models------------'}], columns = df_cols) 
    df9_8 = pd.DataFrame([{'model':'------------bart_base models------------'}], columns = df_cols) 
    df9_9 = pd.DataFrame([{'model':'------------bart_large models------------'}], columns = df_cols) 
    df10 = pd.DataFrame([{'model':'REPORTED BASELINES------------'}], columns = df_cols1) 
    df11 = pd.DataFrame([{'model':'BASELINES------------'}], columns = df_cols1) 
    df12 = pd.DataFrame([{'model':'OUR MODELS------------'}], columns = df_cols1) 
    df12_1 = pd.DataFrame([{'model':'------------bert_base models------------'}], columns = df_cols1) 
    df12_2 = pd.DataFrame([{'model':'------------bert_large models------------'}], columns = df_cols1) 
    df12_3 = pd.DataFrame([{'model':'------------longformer_base models------------'}], columns = df_cols1) 
    df12_4 = pd.DataFrame([{'model':'------------longformer_large models------------'}], columns = df_cols1) 
    df12_5 = pd.DataFrame([{'model':'------------bigbird-pegasus models------------'}], columns = df_cols1) 
    df12_6 = pd.DataFrame([{'model':'------------roberta_base models------------'}], columns = df_cols1) 
    df12_7 = pd.DataFrame([{'model':'------------roberta_large models------------'}], columns = df_cols1) 
    df12_8 = pd.DataFrame([{'model':'------------bart_base models------------'}], columns = df_cols1) 
    df12_9 = pd.DataFrame([{'model':'------------bart_large models------------'}], columns = df_cols1) 
    
    avg_dfs = [df7, df5, df8, df1, df9, df9_1, df3_1,df9_2, df3_2,df9_6, df3_6,df9_7, df3_7,df9_8,df3_8, df9_9,df3_9,df9_3, df3_3,df9_4, df3_4,df9_5, df3_5]
    step_dfs = [df10, df6, df11, df2, df12, df12_1, df4_1,df12_2, df4_2,df12_6, df4_6,df12_7, df4_7,df12_8, df4_8,df12_9, df4_9,df12_3, df4_3,df12_4, df4_4,df12_5, df4_5]
    
    avg_df = pd.concat(avg_dfs)  
    step_df = pd.concat(step_dfs) 
    
    
    result_file = args.models_path+args.dataset+'.eval.results.xlsx'
    avg_sheet = 'avg_rouges'
    step_sheet = 'step_rouges'
    save_eval_results_to_excel(result_file, avg_sheet, avg_df)
    save_eval_results_to_excel(result_file, step_sheet, step_df)
    #make a copy
#    cp_result_file = copy_result_file(result_file)
  
    hs_avg_best_models = check_best_models(df3)
    hs_step_best_models = check_best_models(df4)
    hs_avg_best_bert_base_models = check_best_models(df3_1)
    hs_step_best_bert_base_models = check_best_models(df4_1)
    hs_avg_best_bert_large_models = check_best_models(df3_2)
    hs_step_best_bert_large_models = check_best_models(df4_2)
    
    hs_avg_best_roberta_base_models = check_best_models(df3_6)
    hs_step_best_roberta_base_models = check_best_models(df4_6)
    hs_avg_best_roberta_large_models = check_best_models(df3_7)
    hs_step_best_roberta_large_models = check_best_models(df4_7)
    
    hs_avg_best_bart_base_models = check_best_models(df3_8)
    hs_step_best_bart_base_models = check_best_models(df4_8)
    hs_avg_best_bart_large_models = check_best_models(df3_9)
    hs_step_best_bart_large_models = check_best_models(df4_9)
    
    hs_avg_best_longformer_base_models = check_best_models(df3_3)
    hs_step_best_longformer_base_models = check_best_models(df4_3)
    hs_avg_best_longformer_large_models = check_best_models(df3_4)
    hs_step_best_longformer_large_models = check_best_models(df4_4)
    
    hs_avg_best_bp_models = check_best_models(df3_5)
    hs_step_best_bp_models = check_best_models(df4_5)
    
    BERT_BASE_COLOR='F5F5F5'#grey
    BERT_LARGE_COLOR='FFFFE0'#yellow
    ROBERTA_BASE_COLOR='F5F5F5'#grey
    ROBERTA_LARGE_COLOR='FFFFE0'#yellow
    BART_BASE_COLOR='F5F5F5'#grey
    BART_LARGE_COLOR='FFFFE0'#yellow
    LONGFORMER_BASE_COLOR='ADD8E6'#blue
    LONGFORMER_LARGE_COLOR='FF0000'#red
    BP_COLOR='FFFF33'#yellow
    BORDER_COLOR="696969"
    
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_bert_base_models, color=BERT_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_bert_base_models,color=BERT_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_bert_large_models,color=BERT_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_bert_large_models,color=BERT_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_roberta_base_models, color=ROBERTA_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_roberta_base_models,color=ROBERTA_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_roberta_large_models,color=ROBERTA_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_roberta_large_models,color=ROBERTA_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_bart_base_models, color=BART_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_bart_base_models,color=BART_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_bart_large_models,color=BART_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_bart_large_models,color=BART_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_longformer_base_models, color=LONGFORMER_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_longformer_base_models,color=LONGFORMER_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_longformer_large_models,color=LONGFORMER_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_longformer_large_models,color=LONGFORMER_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_bp_models,color=BP_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_bp_models,color=BP_COLOR, font=True)
    
    
    color_the_best_metric_border(result_file, avg_sheet, hs_avg_best_models, color=BORDER_COLOR)
    color_the_best_metric_border(result_file, step_sheet, hs_step_best_models, color=BORDER_COLOR)
#    color_the_best_metric(cp_result_file, avg_sheet, hs_avg_best_models, color="f0e40a", font=True)
#    color_the_best_metric(cp_result_file, step_sheet, hs_step_best_models, color="f0e40a", font=True)
    
    
    df13,df14 = get_rouges_df(baseline_models2)
    df13_1,df14_1 = get_rouges_df(baseline_bert_base_models)
    df13_2,df14_2 = get_rouges_df(baseline_bert_large_models)
    df13_3,df14_3 = get_rouges_df(baseline_longformer_base_models)
    df13_4,df14_4 = get_rouges_df(baseline_longformer_large_models)
    df13_5,df14_5 = get_rouges_df(baseline_bp_models)
    df13_6,df14_6 = get_rouges_df(baseline_roberta_base_models)
    df13_7,df14_7 = get_rouges_df(baseline_roberta_large_models)
    df13_8,df14_8 = get_rouges_df(baseline_bart_base_models)
    df13_9,df14_9 = get_rouges_df(baseline_bart_large_models)
    
    baseline_avg_best_models = check_best_models(df13)
    baseline_step_best_models = check_best_models(df14)
    baseline_avg_best_bert_base_models = check_best_models(df13_1)
    baseline_step_best_bert_base_models = check_best_models(df14_1)
    baseline_avg_best_bert_large_models = check_best_models(df13_2)
    baseline_step_best_bert_large_models = check_best_models(df14_2)
    
    baseline_avg_best_roberta_base_models = check_best_models(df13_6)
    baseline_step_best_roberta_base_models = check_best_models(df14_6)
    baseline_avg_best_roberta_large_models = check_best_models(df13_7)
    baseline_step_best_roberta_large_models = check_best_models(df14_7)
    
    baseline_avg_best_bart_base_models = check_best_models(df13_8)
    baseline_step_best_bart_base_models = check_best_models(df14_8)
    baseline_avg_best_bart_large_models = check_best_models(df13_9)
    baseline_step_best_bart_large_models = check_best_models(df14_9)
    
    baseline_avg_best_longformer_base_models = check_best_models(df13_3)
    baseline_step_best_longformer_base_models = check_best_models(df14_3)
    baseline_avg_best_longformer_large_models = check_best_models(df13_4)
    baseline_step_best_longformer_large_models = check_best_models(df14_4)
    
    baseline_avg_best_bp_models = check_best_models(df13_5)
    baseline_step_best_bp_models = check_best_models(df14_5)
    
    
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_bert_base_models, color=BERT_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_bert_base_models,color=BERT_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_bert_large_models,color=BERT_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_bert_large_models,color=BERT_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_roberta_base_models, color=ROBERTA_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_roberta_base_models,color=ROBERTA_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_roberta_large_models,color=ROBERTA_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_roberta_large_models,color=ROBERTA_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_bart_base_models, color=BART_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_bart_base_models,color=BART_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_bart_large_models,color=BART_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_bart_large_models,color=BART_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_longformer_base_models, color=LONGFORMER_BASE_COLOR,font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_longformer_base_models,color=LONGFORMER_BASE_COLOR, font=True)
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_longformer_large_models,color=LONGFORMER_LARGE_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_longformer_large_models,color=LONGFORMER_LARGE_COLOR, font=True)
    
    color_the_best_metric(result_file, avg_sheet, baseline_avg_best_bp_models,color=BP_COLOR, font=True)
    color_the_best_metric(result_file, step_sheet, baseline_step_best_bp_models,color=BP_COLOR, font=True)
    
    color_the_best_metric_border(result_file, avg_sheet, baseline_avg_best_models, color=BORDER_COLOR)
    color_the_best_metric_border(result_file, step_sheet, baseline_step_best_models, color=BORDER_COLOR)
#    color_the_best_metric(cp_result_file, avg_sheet, bert_avg_best_models, color="DDDDDD", font=True)
#    color_the_best_metric(cp_result_file, step_sheet, bert_step_best_models, color="DDDDDD", font=True)
    
    mark_best_models(baseline_avg_best_models,avg_df)
    mark_best_models(hs_avg_best_models,avg_df)
    mark_best_models(baseline_step_best_models,step_df)
    mark_best_models(hs_step_best_models,step_df)

    logger.info('avg rouges-------------')
    logger.info(avg_df)
    logger.info('step model rouges------')
    logger.info(step_df)

    logger.info("Generate evaluation results overview...DONE")
    
    #return best step models for plotting summary distribution
    return hs_step_best_models, baseline_step_best_models, hs_avg_best_models, baseline_avg_best_models


                    
def remove_ckp(models,nrm):
    for model in models:
        if model not in nrm:
            
            eval_path = args.models_path + model + '/eval/'
            if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
                
                files = os.listdir(args.models_path + model)
                step_models = [file for file in files if file.endswith('.pt')]
                logger.info("remove %i step models from model %s"%(len(step_models),model))
                for m in step_models:
                    path = args.models_path + model + '/' + m
                    os.remove(path)
                    
            else:
                if not os.path.exists(args.models_path + model+'/DONE'):
                    logger.info("---Training of the model is not finished, skip it-------%s"%(model))
                if not os.path.exists(eval_path+'DONE'):
                    logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
                    
def remove_all_ckp(args,nrm):
    logger.info("=================================================")
    logger.info("Removing all ckp of not best settings...")
    models = sorted(os.listdir(args.models_path))
    models = [model for model in models if model.startswith(args.dataset+'_')]

    remove_ckp(models,nrm)    
    logger.info("Remove all ckp...DONE")
    
def _remove_step_models(models,nrm):
    for model in models:
        if model not in nrm:
            
#            eval_path = args.models_path + model + '/eval/'
#            
#            if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
#                files = os.listdir(eval_path)
#                summ_files = [file for file in files if file.endswith('.gold')]
#                steps = [file.split('.')[1].split('_')[1].split('p')[1] for file in summ_files]
#                
#                files = os.listdir(args.models_path + model)
#                step_models = [file for file in files if file.endswith('.pt')]
#                removed_step_models = [model for model in step_models if not model.split('.')[0].split('_')[-1] in steps]
#                logger.info("remove %i step models from model %s"%(len(removed_step_models),model))
#                for m in removed_step_models:
#                    path = args.models_path + model + '/' + m
#                    os.remove(path)
#                    
#            else:
#                if not os.path.exists(args.models_path + model+'/DONE'):
#                    logger.info("---Training of the model is not finished, skip it-------%s"%(model))
#                if not os.path.exists(eval_path+'DONE'):
#                    logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
            
            eval_path = args.models_path + model + '/eval/'
           
            
            
        
            if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
                 test_rouges_file = eval_path+'test_rouges.json'
                 with open(test_rouges_file) as f:
                     test_rouges = json.load(f)
                 best_model_names = [item[0] for item in test_rouges]
                 best_model_steps = [item.split('_')[-1].replace('.pt','') for item in best_model_names]
                 print(best_model_names)
                 print(best_model_steps)
                 
                 assert 1==2
#                files = os.listdir(eval_path)
#                summ_files = [file for file in files if file.endswith('.gold')]
#                steps = [file.split('.')[1].split('_')[1].split('p')[1] for file in summ_files]
#                
#                files = os.listdir(args.models_path + model)
#                step_models = [file for file in files if file.endswith('.pt')]
#                removed_step_models = [model for model in step_models if not model.split('.')[0].split('_')[-1] in steps]
#                logger.info("remove %i step models from model %s"%(len(removed_step_models),model))
#                for m in removed_step_models:
#                    path = args.models_path + model + '/' + m
#                    os.remove(path)
#                    
            else:
                if not os.path.exists(args.models_path + model+'/DONE'):
                    logger.info("---Training of the model is not finished, skip it-------%s"%(model))
                if not os.path.exists(eval_path+'DONE'):
                    logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))    
    
def remove_step_models(args,nrm):
    logger.info("=================================================")
    logger.info("Removing step models...")
    models = sorted(os.listdir(args.models_path))
    models = [model for model in models if model.startswith(args.dataset+'_')]

    
    _remove_step_models(models,nrm)
        
    logger.info("Remove step models...DONE")


            
            
def plot_val_xent(args):
    logger.info("=================================================")
    logger.info("Plotting validation loss...")
    models = sorted(os.listdir(args.models_path))
    models = [model for model in models if model.startswith(args.dataset+'_')]
    histruct_models = [model for model in models if model.split('_')[1]=='hs']
    
    for model in histruct_models:
        
        eval_path = args.models_path + model + '/eval/'
        
        if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
            
            val_xent_path = eval_path + val_xents_file
            if os.path.exists(val_xent_path):
                with open(val_xent_path, 'r') as f:
                    val_xents = json.load(f)
                    
                    val_xents_dict = {}
                    steps = []
                    for v in val_xents:
                        xent = round(v[0],2)
                    
                        step = int(v[1].split('/')[-1].split('_')[-1].split('.')[0])#
                        
                        steps.append(step)
                        val_xents_dict.update({step:xent})
                    steps.sort()
                    xents = [val_xents_dict[step] for step in steps]
                    
                    png_file = eval_path+'val.xents.png'
                    
                    plt.plot(steps, xents)
                    plt.title('val_xents: '+model)
                    plt.ylabel('val_xents', fontsize='large')
                    plt.xlabel('step', fontsize='large')
                    plt.savefig(png_file, bbox_inches='tight')
                    plt.close()   
            else:
                logger.info("---Validation loss was not saved, skip it-------%s"%(model))
                
           
        else:
            if not os.path.exists(args.models_path + model+'/DONE'):
                logger.info("---Training of the model is not finished, skip it-------%s"%(model))
            if not os.path.exists(eval_path+'DONE'):
                logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
    
    logger.info("Plot validation loss...DONE")
    
def get_prob_dic(modelname, step):
    
    path = args.models_path+modelname+'/eval/'
    file = [file for file in os.listdir(path) if file.endswith('step'+str(step)+'.selectedIdx')][0]
    with open(path+file,'r') as f:
        data = json.load(f)
        flat = [item for sublist in data for item in sublist]  
    prob_dic = []
    for i in range(max(flat)+1):
        prob_dic.append(0)
    for i in set(flat):
        prob_dic[i] = flat.count(i)/len(flat)
    return prob_dic
    
    
def get_best_step_model_prob(best_models):
    flat_list = [item for sublist in list(best_models.values()) for item in sublist]
    dic = ({i:flat_list.count(i) for i in flat_list})
#    dic = ({i:list(best_models.values()).count(i) for i in list(best_models.values())})
    max_v = max(list(dic.values()))
    total = len(metrics)
    best = []
    for model in dic.keys():
        if dic[model]==max_v:
            best.append(model)
            logger.info("Best step model: %s, won on (%i/%i) metrics"%(model,dic[model],total))
    prob_dics = {}        
      
    for model in best:
        name = model[0]
        step = model[1]
        prob_dic = get_prob_dic(name, step)
        prob_dics.update({name+'.step'+str(step):prob_dic})
    return prob_dics
    
       
def plot_best_summ_distribution(args, hs_step_best_models, baseline_step_best_models):
    logger.info("=================================================")
    logger.info("Plotting best summary distribution...")
    prob_dics = {}
    
    oracle = args.dataset+'_oracle'
    
    prob_dics.update({oracle+'.step0': get_prob_dic(oracle, 0)})
    
    flat_list = [item for sublist in list(baseline_step_best_models.values()) for item in sublist]
    dic = ({i:flat_list.count(i) for i in flat_list})
    if len(list(dic.values()))!=0:
        best_baseline_prob_dics = get_best_step_model_prob(baseline_step_best_models)
        prob_dics.update(best_baseline_prob_dics)

    best_hs_prob_dics = get_best_step_model_prob(hs_step_best_models)
    prob_dics.update(best_hs_prob_dics)

    png_file = args.models_path+args.dataset+'.summ.dist.png'
    lens = [len(v) for v in prob_dics.values()]
    max_le = max(lens)
    index = [i for i  in range(max_le)]
    for k in list(prob_dics.keys()):
        if len(prob_dics[k])<max_le:
            prob_dics[k] = prob_dics[k] + [0]*(max_le - len(prob_dics[k]))
       
    df = pd.DataFrame(prob_dics, index=index)
    ax = df.plot.bar(rot=0,figsize=(20,6),title='summary distribution, dataset %s'%(args.dataset))
    ax.set_ylim(0,0.3)
    ax.set_xlabel("sentence position in source text")
    ax.set_ylabel("propotion of selected sentences")
    ax.get_figure().savefig(png_file)
    
    logger.info("Plot best summary distribution...DONE")        
        
     
    


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-generate_eval_results_overview", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-remove_step_models", type=str2bool, nargs='?',const=False,default=False)
    parser.add_argument("-remove_step_models_also_best", type=str2bool, nargs='?',const=False,default=False)
    
    parser.add_argument("-remove_all_ckp_of_not_best", type=str2bool, nargs='?',const=False,default=False)
    parser.add_argument("-plot_val_xent", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-plot_best_summ_distribution", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-plot_summ_distribution", type=str2bool, nargs='?',const=True,default=True)

    
    parser.add_argument("-models_path", default='')
    parser.add_argument("-dataset", default='cnndm', type=str, choices=['cnndm', 'pubmed','arxiv'])
    



    args = parser.parse_args()
    
    init_logger(args.models_path+args.dataset+'.postpro.log')
    logger.info("#################################################")

    os.environ['KMP_DUPLICATE_LIB_OK']='True'
  
    if (args.generate_eval_results_overview):
        hs_step_best_models, baseline_step_best_models,hs_avg_best_models, baseline_avg_best_models = generate_eval_results_overview(args)
        
    if (args.remove_step_models):

        not_removed = []
        if not (args.remove_step_models_also_best):
            flat1 = [item[0] for sublist in list(hs_step_best_models.values()) for item in sublist]
            flat2 = [item[0] for sublist in list(baseline_step_best_models.values()) for item in sublist]
            flat3 = [item[0] for sublist in list(hs_avg_best_models.values()) for item in sublist]
            flat4 = [item[0] for sublist in list(baseline_avg_best_models.values()) for item in sublist]
            not_removed.extend(flat1+flat2+flat3+flat4)
            not_removed=set(not_removed)
            logger.info('Step models of best settings are not removed: '+ ','.join(not_removed))
        else:
            logger.info('Step models of best settings are removed to save space.')
        remove_step_models(args, not_removed)
        
        
    if (args.remove_all_ckp_of_not_best):

        not_removed = []
        
        flat1 = [item[0] for sublist in list(hs_step_best_models.values()) for item in sublist]
        flat2 = [item[0] for sublist in list(baseline_step_best_models.values()) for item in sublist]
        flat3 = [item[0] for sublist in list(hs_avg_best_models.values()) for item in sublist]
        flat4 = [item[0] for sublist in list(baseline_avg_best_models.values()) for item in sublist]
        not_removed.extend(flat1+flat2+flat3+flat4)
        not_removed=set(not_removed)
        logger.info('Step models of best settings are not removed: '+ ','.join(not_removed))
        
        remove_all_ckp(args, not_removed)
    
    if (args.plot_val_xent):
        plot_val_xent(args)
        
    if (args.plot_summ_distribution):
        plot_best_summ_distribution(args, hs_step_best_models, baseline_step_best_models)
        
    
        
    
        

        
        
    
    
  
